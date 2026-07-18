import argparse
import json
import os
import re
import subprocess
import sys
from datetime import datetime
from pathlib import Path

import yaml
import xml.etree.ElementTree as ET
import pandas as pd
from jinja2 import Template
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

def prepare_project_artifact_dirs(project=None):
    # Ensure Reports/ directory exists
    Path("Reports").mkdir(exist_ok=True)

def load_config_details(project_name):
    base_url = "N/A"
    browser = "chrome"
    headless = True
    
    yaml_path = Path(__file__).resolve().parent / ".." / "ALL_PROJECTS_AUTOMATION" / "config" / f"{project_name}.yaml"
    if yaml_path.exists():
        try:
            with open(yaml_path, "r", encoding="utf-8") as f:
                config_data = yaml.safe_load(f)
                if config_data:
                    base_url = config_data.get("base_url", base_url)
                    browser = config_data.get("browser", browser)
                    headless = config_data.get("headless", headless)
        except Exception as exc:
            print(f"Warning: could not load config details: {exc}")
            
    class ConfigModule:
        PROJECT = project_name or "sampark"
        BASE_URL = base_url
        BROWSER = browser
        HEADLESS = headless
        
    return ConfigModule

# ================= CONFIG =================
ROOT = Path.cwd()


def _detect_base_report_dir():
    upper = ROOT / "Reports"
    lower = ROOT / "reports"
    if upper.exists():
        return upper
    if lower.exists():
        return lower
    return upper


def _resolve_report_paths(project_name=None, base_report_dir=None):
    base_dir = Path(base_report_dir) if base_report_dir else _detect_base_report_dir()
    report_dir = base_dir if project_name is None else base_dir / project_name
    report_dir.mkdir(parents=True, exist_ok=True)

    json_path = report_dir / "report.json"
    html_path = report_dir / "report.html"
    excel_path = report_dir / "report.xlsx"
    manual_excel_path = report_dir / "test_report.xlsx"

    return {
        "report_dir": report_dir,
        "json_path": json_path,
        "html_path": html_path,
        "excel_path": excel_path,
        "manual_excel_path": manual_excel_path,
    }


# ================= HTML TEMPLATE =================
HTML_TEMPLATE = """
<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<title>Automation Test Report</title>
<style>
:root {
    --bg-body: #f8fafc;
    --bg-card: #ffffff;
    --border-color: #e2e8f0;
    --text-main: #334155;
    --text-muted: #64748b;
    --text-title: #0f172a;
    --logo-text-main: #2a257c;
    --table-header-bg: #f8fafc;
    --table-row-hover: #f1f5f9;
    --status-pass: #166534;
    --status-fail: #991b1b;
    --status-skipped: #92400e;
    --caution-bg: #fffbeb;
    --caution-border: #fef3c7;
    --caution-text: #92400e;
    --caution-icon: #d97706;
    --footer-meta: #94a3b8;
}

[data-theme="dark"] {
    --bg-body: #0f172a;
    --bg-card: #1e293b;
    --border-color: #334155;
    --text-main: #f8fafc;
    --text-muted: #94a3b8;
    --text-title: #ffffff;
    --logo-text-main: #ffffff;
    --table-header-bg: #0f172a;
    --table-row-hover: #1e293b;
    --status-pass: #4ade80;
    --status-fail: #f87171;
    --status-skipped: #fbbf24;
    --caution-bg: #2d1e10;
    --caution-border: #451a03;
    --caution-text: #fde68a;
    --caution-icon: #f59e0b;
    --footer-meta: #64748b;
}

body {
    font-family: Segoe UI, Arial, sans-serif;
    background: var(--bg-body);
    color: var(--text-main);
    margin: 0;
    padding: 24px;
    transition: background 0.3s, color 0.3s;
}
header {
    display: flex;
    justify-content: space-between;
    align-items: center;
    border-bottom: 1px solid var(--border-color);
    padding-bottom: 20px;
    margin-bottom: 24px;
}
.header-left {
    display: flex;
    flex-direction: column;
}
.header-right {
    display: flex;
    align-items: center;
    gap: 16px;
}
h1 { margin: 0; font-size: 32px; font-weight: 800; color: var(--text-title); }
.meta { color: var(--text-muted); margin: 8px 0 0; font-size: 14px; font-weight: 500; }

.project-info {
    display: flex;
    gap: 20px;
    margin-top: 12px;
    font-size: 13px;
    color: var(--text-main);
    background: var(--bg-card);
    padding: 8px 16px;
    border-radius: 8px;
    border: 1px solid var(--border-color);
    flex-wrap: wrap;
}
.info-item {
    display: flex;
    align-items: center;
    gap: 6px;
}
.info-item strong {
    color: var(--text-muted);
}

.theme-toggle-btn {
    background: var(--bg-card);
    border: 1px solid var(--border-color);
    color: var(--text-main);
    padding: 10px;
    border-radius: 50%;
    cursor: pointer;
    display: flex;
    align-items: center;
    justify-content: center;
    transition: all 0.3s ease;
    box-shadow: 0 4px 12px rgba(0, 0, 0, 0.1);
}
.theme-toggle-btn:hover {
    background: var(--border-color);
    transform: scale(1.05);
}

.logo-container {
    display: flex;
    align-items: center;
    gap: 10px;
    background: var(--bg-card);
    padding: 8px 16px;
    border-radius: 12px;
    border: 1px solid var(--border-color);
    box-shadow: 0 4px 12px rgba(0, 0, 0, 0.1);
}

.cards {
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
    gap: 16px;
    margin: 24px 0;
}
.card {
    padding: 20px;
    border-radius: 16px;
    background: var(--bg-card);
    border: 1px solid var(--border-color);
    box-shadow: 0 8px 24px rgba(0, 0, 0, 0.05);
    transition: all 0.25s cubic-bezier(0.4, 0, 0.2, 1);
}
.card-title { font-size: 13px; color: var(--text-muted); text-transform: uppercase; letter-spacing: 0.08em; }
.value { font-size: 36px; font-weight: 700; margin-top: 10px; color: var(--text-title); }
.small { font-size: 14px; color: var(--text-muted); margin-top: 8px; }
.cards .pass { border-left: 6px solid #22c55e; }
.cards .fail { border-left: 6px solid #ef4444; }
.cards .skip { border-left: 6px solid #f59e0b; }
.card.clickable {
    cursor: pointer;
}
.card.clickable:hover {
    transform: translateY(-4px);
    box-shadow: 0 12px 30px rgba(0, 0, 0, 0.12);
}
.card.clickable.active-filter {
    border-color: #3b82f6;
    box-shadow: 0 0 16px rgba(59, 130, 246, 0.35);
}
.card.pass.clickable.active-filter {
    border-color: #22c55e;
    box-shadow: 0 0 16px rgba(34, 197, 94, 0.35);
}
.card.fail.clickable.active-filter {
    border-color: #ef4444;
    box-shadow: 0 0 16px rgba(239, 68, 68, 0.35);
}
.card.skip.clickable.active-filter {
    border-color: #f59e0b;
    box-shadow: 0 0 16px rgba(245, 158, 11, 0.35);
}
.charts-top-row {
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(380px, 1fr));
    gap: 18px;
    margin-bottom: 18px;
    align-items: start;
}
.charts-bottom-row {
    display: grid;
    grid-template-columns: 1fr;
    gap: 18px;
    margin-bottom: 24px;
    align-items: start;
}
.chart-card {
    padding: 20px;
    border-radius: 16px;
    background: var(--bg-card);
    border: 1px solid var(--border-color);
    box-shadow: 0 8px 24px rgba(0, 0, 0, 0.05);
    min-height: 380px;
    max-height: 420px;
    overflow: hidden;
    color: var(--text-main);
}
.chart-title { font-weight: 700; margin-bottom: 14px; color: var(--text-title); }
.chart-canvas {
    width: 100%;
    height: 320px;
    max-height: 320px;
    display: block;
}
.table-container {
    background: var(--bg-card);
    border: 1px solid var(--border-color);
    border-radius: 16px;
    overflow: auto;
    max-height: 920px;
    box-shadow: 0 8px 24px rgba(0, 0, 0, 0.05);
}
table {
    width: 100%;
    border-collapse: collapse;
    table-layout: fixed;
}
th {
    position: sticky;
    top: 0;
    z-index: 10;
    background: var(--table-header-bg);
    color: var(--text-main);
    text-align: left;
    font-size: 13px;
    font-weight: 700;
    padding: 14px 16px;
    border-bottom: 2px solid var(--border-color);
}
td {
    padding: 14px 16px;
    border-bottom: 1px solid var(--border-color);
    vertical-align: top;
    word-break: break-word;
    font-size: 13px;
    color: var(--text-main);
}
.status-pass { color: var(--status-pass); font-weight: 700; }
.status-fail { color: var(--status-fail); font-weight: 700; }
.status-skipped { color: var(--status-skipped); font-weight: 700; }
.muted { color: var(--text-muted); }

footer {
    margin-top: 48px;
    padding-top: 24px;
    border-top: 1px solid var(--border-color);
}
.footer-caution {
    display: flex;
    align-items: flex-start;
    gap: 12px;
    background: var(--caution-bg);
    border: 1px solid var(--caution-border);
    color: var(--caution-text);
    padding: 16px;
    border-radius: 12px;
    font-size: 13px;
    line-height: 1.6;
    margin-bottom: 20px;
}
.caution-icon {
    margin-top: 2px;
    color: var(--caution-icon);
    flex-shrink: 0;
}
.footer-disclaimer {
    font-size: 12px;
    color: var(--text-muted);
    line-height: 1.5;
    background: var(--bg-card);
    border: 1px solid var(--border-color);
    padding: 14px 16px;
    border-radius: 12px;
    margin-bottom: 24px;
}
.footer-meta {
    display: flex;
    justify-content: space-between;
    align-items: center;
    font-size: 13px;
    color: var(--footer-meta);
    padding: 0 4px;
}
.footer-right {
    font-weight: 600;
    color: var(--text-muted);
}

@media (max-width: 900px) {
    .chart-canvas { height: 240px; }
}
</style>
</head>
<body>
<header>
    <div class="header-left">
        <h1>{{ (project_name or config_project or 'AEPL').upper() }} Automation Test Report</h1>
        <p class="meta">Generated at {{ time }}</p>
        <div class="project-info">
            <div class="info-item">
                <strong>Project:</strong>
                <span>{{ (project_name or config_project or 'LCT').upper() }}</span>
            </div>
            <div class="info-item">
                <strong>Target URL:</strong>
                <a href="{{ base_url }}" target="_blank" style="color: #3b82f6; text-decoration: none; font-weight: 600;">{{ base_url }}</a>
            </div>
            <div class="info-item">
                <strong>Browser:</strong>
                <span>{{ browser }} ({{ 'Headless' if headless else 'Headed' }})</span>
            </div>
            <div class="info-item">
                <strong>Test Suite:</strong>
                <span>{{ markers }}</span>
            </div>
            <div class="info-item">
                <strong>Execution Duration:</strong>
                <span>{{ total_duration }}s</span>
            </div>
        </div>
    </div>
    <div class="header-right">
        <button id="themeToggle" class="theme-toggle-btn" aria-label="Toggle Theme">
            <svg class="sun-icon" width="20" height="20" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round">
                <circle cx="12" cy="12" r="5"></circle>
                <line x1="12" y1="1" x2="12" y2="3"></line>
                <line x1="12" y1="21" x2="12" y2="23"></line>
                <line x1="4.22" y1="4.22" x2="5.64" y2="5.64"></line>
                <line x1="18.36" y1="18.36" x2="19.78" y2="19.78"></line>
                <line x1="1" y1="12" x2="3" y2="12"></line>
                <line x1="21" y1="12" x2="23" y2="12"></line>
                <line x1="4.22" y1="19.78" x2="5.64" y2="18.36"></line>
                <line x1="18.36" y1="5.64" x2="19.78" y2="4.22"></line>
            </svg>
            <svg class="moon-icon" width="20" height="20" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round" style="display: none;">
                <path d="M21 12.79A9 9 0 1 1 11.21 3 7 7 0 0 0 21 12.79z"></path>
            </svg>
        </button>
        <div class="logo-container" style="display: flex; align-items: center; justify-content: center; width: 180px; height: 120px; padding: 8px; overflow: hidden; border-radius: 16px;">
            {% if logo_base64 %}
            <img src="data:image/png;base64,{{ logo_base64 }}" alt="Accolade Logo" style="width: 100%; height: 100%; object-fit: contain;">
            {% else %}
            <span style="font-weight: 800; font-size: 16px; color: var(--text-title);">AEPL</span>
            {% endif %}
        </div>
    </div>
</header>
<div class="cards">
    <div class="card clickable" id="totalCard">
        <div class="card-title">Total Tests</div>
        <div class="value">{{ total }}</div>
        <div class="small">Overall execution count</div>
    </div>
    <div class="card pass clickable" id="passCard">
        <div class="card-title">Passed</div>
        <div class="value">{{ passed }}</div>
        <div class="small">Successful assertions</div>
    </div>
    <div class="card fail clickable" id="failCard">
        <div class="card-title">Failed</div>
        <div class="value">{{ failed }}</div>
        <div class="small">Requires attention</div>
    </div>
    <div class="card skip clickable" id="skipCard">
        <div class="card-title">Skipped</div>
        <div class="value">{{ skipped }}</div>
        <div class="small">Conditional or blocked tests</div>
    </div>
    <div class="card">
        <div class="card-title">Avg Duration</div>
        <div class="value">{{ average_duration }}</div>
        <div class="small">Seconds per test</div>
    </div>
</div>

<div class="charts-top-row">
    <div class="chart-card">
        <div class="chart-title">Test Status Distribution</div>
        <canvas id="statusChart" class="chart-canvas"></canvas>
    </div>
    <div class="chart-card">
        <div class="chart-title">Top Slow Tests</div>
        <canvas id="slowChart" class="chart-canvas"></canvas>
    </div>
</div>

<div class="charts-bottom-row">
    <div class="chart-card" style="max-height: none; min-height: 400px;">
        <div class="chart-title">Test Duration Trend</div>
        <canvas id="durationChart" class="chart-canvas"></canvas>
    </div>
</div>

<div class="table-container">
<table>
<tr>
    <th style="width: 24%;">Test Case Name</th>
    <th style="width: 20%;">Expected</th>
    <th style="width: 20%;">Actual</th>
    <th style="width: 10%;">Result</th>
    <th style="width: 10%;">Duration</th>
    <th style="width: 16%;">Message</th>
</tr>
{% for t in tests %}
<tr class="test-row" data-status="{{ t.status }}">
    <td>{{ t.name }}</td>
    <td>{{ t.expected }}</td>
    <td>{{ t.actual }}</td>
    <td class="status-{{ t.status }}">{{ t.status|upper }}</td>
    <td>{{ t.duration }}{% if t.duration != "" %}s{% endif %}</td>
    <td class="muted">{{ t.message }}</td>
</tr>
{% endfor %}
</table>
</div>

<footer>
    <div class="footer-caution">
        <svg class="caution-icon" width="20" height="20" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round">
            <path d="M10.29 3.86L1.82 18a2 2 0 0 0 1.71 3h16.94a2 2 0 0 0 1.71-3L13.71 3.86a2 2 0 0 0-3.42 0z"/>
            <line x1="12" y1="9" x2="12" y2="13"/>
            <line x1="12" y1="17" x2="12.01" y2="17"/>
        </svg>
        <span><strong>CAUTION:</strong> This is a secure system diagnostic report. The outcomes, logs, and metadata contained within are confidential and intended exclusively for internal review and validation by authorized team members. Unauthorized duplication, modification, or distribution is strictly prohibited.</span>
    </div>
    <div class="footer-disclaimer">
        <strong>Disclaimer:</strong> This automated test execution report is generated dynamically by the Accolade Electronics QA Testing Pipeline. The testing outcomes, metrics, and logs present a point-in-time assessment of the application's functional sanity under controlled execution parameters. Actual production environment behavior may vary depending on environmental factors, network latency, database states, or external integrations.
    </div>
    <div class="footer-meta">
        <span>AEPL Automation Diagnostic System &copy; {{ current_year }}. All rights reserved.</span>
        <span>Run by: <strong>{{ run_by }}</strong> | Environment: <strong>QA</strong></span>
    </div>
</footer>

<script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
<script>
const statusData = {
    labels: ['Passed', 'Failed', 'Skipped'],
    datasets: [{
        data: [{{ passed }}, {{ failed }}, {{ skipped }}],
        backgroundColor: ['#22c55e', '#ef4444', '#f59e0b'],
        borderWidth: 0
    }]
};
const statusChart = new Chart(document.getElementById('statusChart'), {
    type: 'doughnut',
    data: statusData,
    options: {
        responsive: true,
        maintainAspectRatio: false,
        plugins: {
            legend: {
                position: 'bottom',
                labels: {
                    usePointStyle: true,
                    color: '#cbd5e1'
                }
            }
        }
    }
});

const durationData = {
    labels: [{% for t in tests %}'{{ t.name | replace("'", "\'") }}'{% if not loop.last %}, {% endif %}{% endfor %}],
    datasets: [{
        label: 'Duration (s)',
        data: [{% for t in tests %}{{ t.duration or 0 }}{% if not loop.last %}, {% endif %}{% endfor %}],
        borderColor: '#2563eb',
        backgroundColor: 'rgba(59, 130, 246, 0.15)',
        fill: true,
        tension: 0.25,
        pointRadius: 3,
        borderWidth: 2
    }]
};
const durationChart = new Chart(document.getElementById('durationChart'), {
    type: 'line',
    data: durationData,
    options: {
        responsive: true,
        maintainAspectRatio: false,
        scales: {
            x: {
                ticks: {
                    color: '#94a3b8',
                    autoSkip: true,
                    maxTicksLimit: 25,
                    maxRotation: 0,
                    minRotation: 0,
                    callback: function(value, index, values) {
                        return index + 1;
                    }
                },
                grid: {
                    color: '#334155'
                }
            },
            y: {
                beginAtZero: true,
                max: {{ y_max }},
                ticks: {
                    color: '#94a3b8'
                },
                grid: {
                    color: '#334155'
                },
                title: {
                    display: true,
                    text: 'Duration (s)',
                    color: '#cbd5e1',
                    font: { weight: 'bold' }
                }
            }
        },
        plugins: {
            legend: {
                labels: {
                    color: '#cbd5e1'
                }
            },
            tooltip: {
                callbacks: {
                    title: function(context) {
                        return context[0].label;
                    }
                }
            }
        }
    }
});

const slowTests = [
{% for t in slow_tests %}
    { name: '{{ t.name | replace("'", "\'") }}', duration: {{ t.duration }} }{% if not loop.last %}, {% endif %}
{% endfor %}
];
const slowData = {
    labels: slowTests.map(item => item.name),
    datasets: [{
        label: 'Duration (s)',
        data: slowTests.map(item => item.duration),
        backgroundColor: '#f97316'
    }]
};
const slowChart = new Chart(document.getElementById('slowChart'), {
    type: 'bar',
    data: slowData,
    options: {
        indexAxis: 'y',
        responsive: true,
        maintainAspectRatio: false,
        scales: {
            x: {
                beginAtZero: true,
                ticks: {
                    color: '#94a3b8'
                },
                grid: {
                    color: '#334155'
                },
                title: {
                    display: true,
                    text: 'Duration (s)',
                    color: '#cbd5e1',
                    font: { weight: 'bold' }
                }
            },
            y: {
                ticks: {
                    color: '#94a3b8',
                    autoSkip: true,
                    callback: function(value, index, values) {
                        const label = this.getLabelForValue(value);
                        return label.length > 20 ? label.substring(0, 20) + '...' : label;
                    }
                },
                grid: {
                    color: '#334155'
                }
            }
        },
        plugins: {
            legend: { display: false },
            tooltip: {
                callbacks: {
                    title: function(context) {
                        return context[0].label;
                    }
                }
            }
        }
    }
});

const themeToggle = document.getElementById('themeToggle');
const sunIcon = themeToggle.querySelector('.sun-icon');
const moonIcon = themeToggle.querySelector('.moon-icon');

function updateChartTheme(theme) {
    const textColor = theme === 'light' ? '#334155' : '#cbd5e1';
    const gridColor = theme === 'light' ? '#e2e8f0' : '#334155';

    // Update duration chart
    durationChart.options.scales.x.ticks.color = textColor;
    durationChart.options.scales.x.grid.color = gridColor;
    durationChart.options.scales.y.ticks.color = textColor;
    durationChart.options.scales.y.grid.color = gridColor;
    durationChart.options.scales.y.title.color = textColor;
    durationChart.options.plugins.legend.labels.color = textColor;
    durationChart.update();

    // Update slow chart
    slowChart.options.scales.x.ticks.color = textColor;
    slowChart.options.scales.x.grid.color = gridColor;
    slowChart.options.scales.x.title.color = textColor;
    slowChart.options.scales.y.ticks.color = textColor;
    slowChart.options.scales.y.grid.color = gridColor;
    slowChart.update();

    // Update status chart
    statusChart.options.plugins.legend.labels.color = textColor;
    statusChart.update();
}

themeToggle.addEventListener('click', () => {
    const currentTheme = document.documentElement.getAttribute('data-theme') || 'light';
    const newTheme = currentTheme === 'light' ? 'dark' : 'light';
    
    document.documentElement.setAttribute('data-theme', newTheme);
    
    if (newTheme === 'light') {
        sunIcon.style.display = 'none';
        moonIcon.style.display = 'block';
    } else {
        sunIcon.style.display = 'block';
        moonIcon.style.display = 'none';
    }
    
    updateChartTheme(newTheme);
    localStorage.setItem('theme', newTheme);
});

// Load preferred theme on startup
const savedTheme = localStorage.getItem('theme') || 'light';
if (savedTheme === 'dark') {
    document.documentElement.setAttribute('data-theme', 'dark');
    sunIcon.style.display = 'block';
    moonIcon.style.display = 'none';
    window.addEventListener('load', () => updateChartTheme('dark'));
} else {
    document.documentElement.setAttribute('data-theme', 'light');
    sunIcon.style.display = 'none';
    moonIcon.style.display = 'block';
    window.addEventListener('load', () => updateChartTheme('light'));
}

// Interactive status filtering
const filterCards = {
    total: document.getElementById('totalCard'),
    pass: document.getElementById('passCard'),
    fail: document.getElementById('failCard'),
    skip: document.getElementById('skipCard')
};
const testRows = document.querySelectorAll('.test-row');
let currentActiveFilter = 'total';

function applyStatusFilter(status) {
    // Remove active filter class from all clickable cards
    Object.values(filterCards).forEach(card => {
        if (card) card.classList.remove('active-filter');
    });

    currentActiveFilter = status;
    if (filterCards[status]) {
        filterCards[status].classList.add('active-filter');
    }

    testRows.forEach(row => {
        const rowStatus = row.getAttribute('data-status');
        if (status === 'total') {
            row.style.display = '';
        } else if (status === 'pass' && (rowStatus === 'pass' || rowStatus === 'passed')) {
            row.style.display = '';
        } else if (status === 'fail' && (rowStatus === 'fail' || rowStatus === 'failed')) {
            row.style.display = '';
        } else if (status === 'skip' && (rowStatus === 'skip' || rowStatus === 'skipped')) {
            row.style.display = '';
        } else {
            row.style.display = 'none';
        }
    });

    // Smooth scroll to the table container
    const tableContainer = document.querySelector('.table-container');
    if (tableContainer) {
        tableContainer.scrollIntoView({ behavior: 'smooth', block: 'start' });
    }
}

if (filterCards.total) filterCards.total.addEventListener('click', () => applyStatusFilter('total'));
if (filterCards.pass) filterCards.pass.addEventListener('click', () => applyStatusFilter('pass'));
if (filterCards.fail) filterCards.fail.addEventListener('click', () => applyStatusFilter('fail'));
if (filterCards.skip) filterCards.skip.addEventListener('click', () => applyStatusFilter('skip'));

// Initialize active state
if (filterCards.total) filterCards.total.classList.add('active-filter');
</script>
</body>
</html>
"""



def _clean(value, limit=500):
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    value = str(value).replace("\r", " ").replace("\n", " ").strip()
    value = re.sub(r"\s+", " ", value)
    return value[:limit]


def _test_case_name(nodeid):
    if not nodeid:
        return ""
    return nodeid.split("::")[-1]


def _normalize_status(value):
    value = str(value or "").strip().lower()
    if value in {"passed", "pass"}:
        return "pass"
    if value in {"failed", "fail"}:
        return "fail"
    return "skipped"


def _duration_seconds(test_result):
    duration = test_result.get("duration")
    if duration is not None:
        return round(duration, 2)

    total = 0
    found = False
    for section in ("setup", "call", "teardown"):
        section_data = test_result.get(section) or {}
        section_duration = section_data.get("duration")
        if section_duration is not None:
            total += section_duration
            found = True

    return round(total, 2) if found else ""


def _property_map(test_result):
    properties = {}
    raw_properties = (
        test_result.get("user_properties") or test_result.get("properties") or []
    )

    if isinstance(raw_properties, dict):
        return {str(key).lower(): value for key, value in raw_properties.items()}

    for item in raw_properties:
        if isinstance(item, dict):
            name = item.get("name") or item.get("key")
            value = item.get("value")
            if name:
                properties[str(name).lower()] = value
            else:
                for key, value in item.items():
                    properties[str(key).lower()] = value
        elif isinstance(item, (list, tuple)) and len(item) >= 2:
            name, value = item[0], item[1]
            if name:
                properties[str(name).lower()] = value
        else:
            continue

    return properties


def _load_manual_results(manual_excel_path):
    """Read richer expected/actual rows written by utils.excel_report.write_result."""
    manual_path = Path(manual_excel_path)
    if not manual_path.exists():
        return {}

    try:
        df = pd.read_excel(manual_path)
    except Exception as exc:
        print(f"Warning: Could not read manual Excel report {manual_path}: {exc}")
        return {}

    required = {"Test Name", "Expected", "Actual", "Status"}
    missing = required - set(df.columns)
    if missing:
        print(f"Warning: Manual report missing columns: {missing}")
        return {}

    results = {}
    for _, row in df.iterrows():
        test_name = _clean(row.get("Test Name"))
        if not test_name:
            continue
        results[test_name] = {
            "expected": _clean(row.get("Expected")),
            "actual": _clean(row.get("Actual")),
            "status": _normalize_status(row.get("Status")),
            "message": _clean(row.get("Error"), 300),
        }
    return results


def _extract_expected_actual(longrepr, outcome):
    """Best-effort extraction for tests that do not call write_result()."""
    text = _clean(longrepr, 2000)

    if not text:
        return "Not recorded by test", "Not recorded by test", ""

    patterns = [
        r"Expected:\s*(?P<expected>.*?)[,;]\s*(?:Got|Actual):\s*(?P<actual>.*)",
        r"Expected\s+'(?P<expected>.*?)'\s*,?\s*(?:but\s+)?got\s+'(?P<actual>.*?)'",
        r"Expected\s+(?P<expected>.*?)\s*,\s*got\s+(?P<actual>.*)",
        r"expected\s+(?P<expected>.*?)\s+but\s+got\s+(?P<actual>.*)",
        r"Expected\s+URL\s+to\s+be\s+'(?P<expected>.*?)',\s+got\s+(?P<actual>.*)",
        r"Expected\s+page\s+title\s+'(?P<expected>.*?)',\s+but\s+got\s+'(?P<actual>.*?)'",
        r"Expected\s+title\s+'(?P<expected>.*?)',\s+but\s+got\s+'(?P<actual>.*?)'",
    ]

    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            return (
                _clean(match.group("expected"), 500),
                _clean(match.group("actual"), 500),
                text[:300],
            )

    if "AssertionError:" in text:
        message = text.split("AssertionError:", 1)[-1].strip()
        return "Assertion condition should pass", _clean(message, 500), text[:300]

    return "Not recorded by test", _clean(text, 500), text[:300]


def _build_test_rows(data, manual_excel_path):
    manual_results = _load_manual_results(manual_excel_path)
    tests = []
    seen = set()
    counts = {"passed": 0, "failed": 0, "skipped": 0}

    for t in data.get("tests", []):
        nodeid = t.get("nodeid", "")
        test_name = _test_case_name(nodeid)
        status = _normalize_status(t.get("outcome", ""))
        properties = _property_map(t)

        # 🔹 PRIORITY 1: Check properties (set by report_case fixture or pytest hook)
        properties_expected = _clean(properties.get("expected"))
        properties_actual = _clean(properties.get("actual"))
        properties_message = _clean(properties.get("message"), 300)

        # 🔹 PRIORITY 2: Extract from error/longrepr only if properties don't have expected/actual
        if properties_expected or properties_actual:
            # Properties has explicit values - use them
            expected = properties_expected
            actual = properties_actual
            message = properties_message or ""
        else:
            # Fall back to extracting from error message
            expected, actual, message = _extract_expected_actual(
                t.get("longrepr", ""), t.get("outcome", "")
            )
            # Override message with property if available
            message = properties_message or message

        # 🔹 PRIORITY 3: Update status from properties if available
        status = _normalize_status(
            properties.get("result") or properties.get("status") or status
        )

        # If report_case stored properties but they are empty strings,
        # preserve the default expected/actual values from the test itself.
        if not expected and properties.get("expected") == "":
            expected = "Not recorded by test"
        if not actual and properties.get("actual") == "":
            actual = "Not recorded by test"

        # 🔹 PRIORITY 4: Manual results (from write_result) override everything
        manual = manual_results.get(test_name)
        if manual:
            expected = manual["expected"] or expected
            actual = manual["actual"] or actual
            status = manual["status"] or status
            message = manual["message"] or message

        counts[{"pass": "passed", "fail": "failed"}.get(status, "skipped")] += 1
        seen.add(test_name)

        # [DEBUG] logging for "Not recorded by test" issues
        if "Not recorded by test" in (expected or actual):
            print(
                f"[DEBUG] Test '{test_name}' has 'Not recorded by test' in expected/actual"
            )
            print(f"   - has properties: {bool(properties)}")
            print(f"   - has manual: {bool(manual)}")
            print(f"   - longrepr length: {len(str(t.get('longrepr', '')))}")

        tests.append(
            {
                "name": test_name,
                "nodeid": nodeid,
                "expected": expected,
                "actual": actual,
                "status": status,
                "duration": _duration_seconds(t),
                "message": _clean(message, 300),
            }
        )

    for test_name, manual in manual_results.items():
        if test_name in seen:
            continue
        status = manual["status"]
        counts[{"pass": "passed", "fail": "failed"}.get(status, "skipped")] += 1
        tests.append(
            {
                "name": test_name,
                "nodeid": test_name,
                "expected": manual["expected"],
                "actual": manual["actual"],
                "status": status,
                "duration": "",
                "message": manual["message"],
            }
        )

    return tests, counts, len(tests)


# ================= STEP 1: RUN PYTEST (MAVEN) =================
def run_pytest(json_path, project_name=None, markers=None):
    report_dir = json_path.parent
    report_dir.mkdir(parents=True, exist_ok=True)

    cmd = ["mvn", "test"]
    
    if project_name:
        cmd.append(f"-Dproject={project_name}")
    if markers:
        cmd.append(f"-Dgroups={markers}")
        
    test_arg = os.environ.get("MAVEN_TEST")
    if test_arg:
        cmd.append(f"-Dtest={test_arg}")
        
    print("=" * 60)
    print("RUNNING JAVA SELENIUM TESTS (MAVEN)")
    print("=" * 60)
    print("Command:", " ".join(cmd))
    print("-" * 60)

    result = subprocess.run(" ".join(cmd), shell=True, cwd=ROOT)
    return result.returncode


# ================= STEP 2: PROCESS TESTNG RESULTS =================
def load_all_excel_results():
    results = {}
    excel_dir = Path("Results/test-results")
    if not excel_dir.exists():
        return results
        
    for file in excel_dir.glob("*.xlsx"):
        try:
            df = pd.read_excel(file)
            required = {"Test Name", "Expected", "Actual", "Status"}
            if not required.issubset(df.columns):
                continue
                
            for _, row in df.iterrows():
                test_name = str(row.get("Test Name", "")).strip()
                if not test_name:
                    continue
                results[test_name.lower()] = {
                    "expected": str(row.get("Expected", "")).strip(),
                    "actual": str(row.get("Actual", "")).strip(),
                    "status": str(row.get("Status", "")).strip().lower(),
                    "message": str(row.get("Error", "")).strip() if "Error" in df.columns else ""
                }
        except Exception as exc:
            pass
            
    return results

def process_json(json_path, manual_excel_path):
    xml_path = Path("target/surefire-reports/testng-results.xml")
    tests = []
    counts = {"passed": 0, "failed": 0, "skipped": 0}
    
    manual_results = load_all_excel_results()
    
    if xml_path.exists():
        try:
            tree = ET.parse(xml_path)
            root = tree.getroot()
            
            for class_elem in root.iter("class"):
                class_name = class_elem.attrib.get("name", "")
                class_simple = class_name.split(".")[-1]
                
                for test_method in class_elem.iter("test-method"):
                    if test_method.attrib.get("is-config") == "true":
                        continue
                        
                    name = test_method.attrib.get("name")
                    full_name = f"{class_simple}.{name}"
                    
                    status = test_method.attrib.get("status", "").lower()
                    if status == "pass":
                        status = "passed"
                    elif status == "fail":
                        status = "failed"
                    elif status == "skip":
                        status = "skipped"
                        
                    duration_ms = float(test_method.attrib.get("duration-ms", 0))
                    duration_s = round(duration_ms / 1000.0, 2)
                    
                    message = ""
                    exception = test_method.find("exception")
                    if exception is not None:
                        message_elem = exception.find("message")
                        if message_elem is not None and message_elem.text:
                            message = message_elem.text.strip()
                        else:
                            message = exception.attrib.get("class", "Exception")
                            
                    expected = "Assertion condition should pass"
                    actual = "Pass" if status == "passed" else "Fail"
                    
                    # Check matching
                    match = None
                    for key in manual_results:
                        if key in full_name.lower() or full_name.lower() in key:
                            match = manual_results[key]
                            break
                            
                    if match:
                        expected = match["expected"]
                        actual = match["actual"]
                        if match["message"]:
                            message = match["message"]
                    else:
                        if status == "failed" and message:
                            if "Expected:" in message and "got:" in message:
                                try:
                                    parts = message.split("Expected:")[-1].split("but got:")
                                    expected = parts[0].strip()
                                    actual = parts[1].strip()
                                except:
                                    pass
                                    
                    tests.append({
                        "name": full_name,
                        "nodeid": f"{class_name}.{name}",
                        "expected": expected,
                        "actual": actual,
                        "status": status,
                        "duration": duration_s,
                        "message": message
                    })
                    counts[status] += 1
                
        except Exception as exc:
            print(f"Error parsing TestNG results: {exc}")
            
    total = len(tests)
    return tests, counts, total, {"tests": tests}


def _average_duration(tests):
    durations = [
        t["duration"] for t in tests if isinstance(t["duration"], (int, float))
    ]
    if not durations:
        return 0
    return round(sum(durations) / len(durations), 2)


def _top_slow_tests(tests, limit=5):
    slow_tests = sorted(
        [t for t in tests if isinstance(t["duration"], (int, float))],
        key=lambda item: item["duration"],
        reverse=True,
    )
    return slow_tests[:limit]


# ================= STEP 3: HTML =================
def generate_html(tests, counts, total, html_path, project_name=None, markers=None):
    try:
        import config.config as config_module
    except ModuleNotFoundError:
        config_module = load_config_details(project_name)
    import getpass
    import base64
    template = Template(HTML_TEMPLATE)

    logo_base64 = ""
    logo_path = Path(__file__).resolve().parent / "report_assets" / "accolade_logo.png"
    if logo_path.exists():
        try:
            logo_base64 = base64.b64encode(logo_path.read_bytes()).decode("utf-8")
        except Exception:
            pass

    run_by = os.getenv("EXECUTION_USER") or os.getenv("GITHUB_ACTOR") or getpass.getuser()
    total_duration = round(sum(t["duration"] for t in tests if isinstance(t["duration"], (int, float))), 2)

    durations = [t["duration"] for t in tests if isinstance(t["duration"], (int, float)) and t["duration"] > 0]
    if durations:
        sorted_dur = sorted(durations)
        idx = min(len(sorted_dur) - 1, int(len(sorted_dur) * 0.95))
        p95 = sorted_dur[idx]
        y_max = int(max(10, round(p95 * 2.0)))
    else:
        y_max = 50

    html = template.render(
        time=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        total=total,
        passed=counts["passed"],
        failed=counts["failed"],
        skipped=counts["skipped"],
        average_duration=_average_duration(tests),
        total_duration=total_duration,
        logo_base64=logo_base64,
        y_max=y_max,
        tests=tests,
        slow_tests=_top_slow_tests(tests),
        project_name=project_name,
        config_project=getattr(config_module, "PROJECT", "lct"),
        base_url=getattr(config_module, "BASE_URL", "N/A"),
        browser=getattr(config_module, "BROWSER", "chromium"),
        headless=getattr(config_module, "HEADLESS", True),
        markers=markers or "All",
        run_by=run_by,
        current_year=datetime.now().year
    )

    html_path.parent.mkdir(parents=True, exist_ok=True)
    html_path.write_text(html, encoding="utf-8")
    print("HTML generated:", html_path)


# ================= STEP 4: EXCEL =================
def generate_excel(tests, excel_path, project_name=None):
    rows = [
        {
            "Test Case Name": t["name"],
            "Expected": t["expected"],
            "Actual": t["actual"],
            "Result": t["status"].upper(),
            "Duration (s)": t["duration"],
            "Message": t["message"],
        }
        for t in tests
    ]

    df = pd.DataFrame(
        rows,
        columns=[
            "Test Case Name",
            "Expected",
            "Actual",
            "Result",
            "Duration (s)",
            "Message",
        ],
    )

    excel_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Test Results", startrow=5)

    _style_excel(excel_path, project_name)

    print("Excel generated:", excel_path)
    print(f"Total rows: {len(rows)}")


def _style_excel(excel_path, project_name=None):
    import getpass
    import os
    from datetime import datetime
    try:
        import config.config as config_module
    except ModuleNotFoundError:
        config_module = load_config_details(project_name)
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.drawing.image import Image

    wb = load_workbook(excel_path)
    ws = wb["Test Results"]

    # 1. Insert Logo
    ws.merge_cells("A1:B4")
    logo_path = Path(__file__).resolve().parent / "report_assets" / "accolade_logo.png"
    if logo_path.exists():
        img = Image(str(logo_path))
        img.width = 130
        img.height = 130
        ws.add_image(img, "A1")

    # 2. Write Metadata
    run_by = os.getenv("EXECUTION_USER") or os.getenv("GITHUB_ACTOR") or getpass.getuser()
    base_url = getattr(config_module, "BASE_URL", "N/A")
    browser = getattr(config_module, "BROWSER", "chromium")

    proj_display = (project_name or getattr(config_module, "PROJECT", "lct")).upper()
    execution_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    metadata = [
        ("Project Name:", f"{proj_display}"),
        ("Target URL:", base_url),
        ("Browser:", browser.upper()),
        ("Test Executed by:", run_by)
    ]

    label_font = Font(name="Segoe UI", size=10, bold=True, color="475569")
    value_font = Font(name="Segoe UI", size=10, color="0f172a")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="ffffff")
    data_font = Font(name="Segoe UI", size=10, color="334155")

    thin_border = Border(
        left=Side(style='thin', color='cbd5e1'),
        right=Side(style='thin', color='cbd5e1'),
        top=Side(style='thin', color='cbd5e1'),
        bottom=Side(style='thin', color='cbd5e1')
    )

    for r_idx, (l1, v1) in enumerate(metadata, start=1):
        c_cell = ws.cell(row=r_idx, column=3, value=l1)
        c_cell.font = label_font
        c_cell.alignment = Alignment(horizontal="right")

        d_cell = ws.cell(row=r_idx, column=4, value=v1)
        d_cell.font = value_font
        d_cell.alignment = Alignment(horizontal="left")

    for r in range(1, 5):
        ws.row_dimensions[r].height = 28

    ws.row_dimensions[5].height = 15

    header_fill = PatternFill("solid", fgColor="1e293b")
    for cell in ws[6]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = thin_border

    ws.row_dimensions[6].height = 26

    pass_fill = PatternFill("solid", fgColor="d1fae5")
    fail_fill = PatternFill("solid", fgColor="fee2e2")
    skip_fill = PatternFill("solid", fgColor="fef3c7")

    result_col = None
    for cell in ws[6]:
        if cell.value == "Result":
            result_col = cell.column
            break

    for row in range(7, ws.max_row + 1):
        ws.row_dimensions[row].height = 20
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

        if result_col:
            cell = ws.cell(row=row, column=result_col)
            result = str(cell.value or "").upper()
            if result == "PASS":
                cell.fill = pass_fill
                cell.font = Font(name="Segoe UI", size=10, bold=True, color="065f46")
            elif result == "FAIL":
                cell.fill = fail_fill
                cell.font = Font(name="Segoe UI", size=10, bold=True, color="991b1b")
            else:
                cell.fill = skip_fill
                cell.font = Font(name="Segoe UI", size=10, bold=True, color="92400e")

    widths = {
        "A": 36,
        "B": 48,
        "C": 30,
        "D": 36,
        "E": 24,
        "F": 55,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    ws.freeze_panes = "A7"
    ws.auto_filter.ref = f"A6:F{ws.max_row}"
    wb.save(excel_path)


# ================= MAIN =================
def _resolve_project_name(cli_project=None):
    if cli_project:
        return cli_project.strip()

    env_project = os.getenv("PROJECT") or os.getenv("PYTEST_PROJECT")
    if env_project:
        return env_project.strip()

    return None


def _resolve_markers(cli_markers=None):
    if cli_markers:
        cleaned = []
        for item in cli_markers:
            parts = item.split(",")
            for p in parts:
                p_clean = p.strip()
                if p_clean:
                    cleaned.append(p_clean)
        raw_joined = " ".join(cli_markers).lower()
        has_logical = any(w in raw_joined for w in [" or ", " and ", " not "])
        if has_logical:
            return " ".join(cleaned)
        return " or ".join(cleaned)

    if os.getenv("CI"):
        return "smoke"

    return None


def parse_args():
    parser = argparse.ArgumentParser()
    parser.add_argument("--project", help="Project name to run.")
    parser.add_argument("--projects", help="Comma-separated project names.")
    parser.add_argument(
        "--marker",
        "--markers",
        dest="markers",
        nargs="+",
        help="Pytest marker expression to select tests, e.g. smoke, ui.",
    )
    parser.add_argument(
        "--test",
        help="Test class or method to run, e.g. LoginPageTest or LoginPageTest#loginSuccess.",
    )
    parser.add_argument(
        "--report-dir",
        help="Optional base report directory. Defaults to Reports/ or reports/.",
    )
    parser.add_argument(
        "--skip-pytest",
        action="store_true",
        help="Skip running pytest and only generate reports from existing JSON.",
    )
    return parser.parse_args()


def _run_report_for_target(project_name, base_report_dir, skip_pytest, markers=None):
    if project_name:
        os.environ["PROJECT"] = project_name
    config_module = load_config_details(project_name)

    prepare_project_artifact_dirs(project=project_name)

    report_paths = _resolve_report_paths(project_name, base_report_dir)
    report_dir = report_paths["report_dir"]
    json_path = report_paths["json_path"]
    html_path = report_paths["html_path"]
    excel_path = report_paths["excel_path"]

    print("=" * 60)
    print(f"TEST REPORT GENERATOR: {project_name or 'default'}")
    print("=" * 60)
    print(f"Working directory: {ROOT}")
    print(f"Report directory: {report_dir}")
    print(f"JSON path: {json_path}")
    print(f"HTML path: {html_path}")
    print(f"Excel path: {excel_path}")
    print(f"Manual expected/actual path: {report_paths['manual_excel_path']}")
    print(f"JSON exists: {json_path.exists()}")
    print("-" * 60)

    if skip_pytest:
        print("Skipping pytest (--skip-pytest specified)")
        exit_code = 0
    else:
        exit_code = run_pytest(json_path, project_name, markers)

    try:
        tests, counts, total, data = process_json(
            json_path, report_paths["manual_excel_path"]
        )
        print("Generating HTML report...")
        generate_html(tests, counts, total, html_path, project_name, markers)

        print("Generating Excel report...")
        generate_excel(tests, excel_path, project_name)

        print("\n" + "=" * 60)
        print("ALL REPORTS GENERATED SUCCESSFULLY!")
        print("=" * 60)
        print(f"HTML:  {html_path} (exists: {html_path.exists()})")
        print(f"Excel: {excel_path} (exists: {excel_path.exists()})")
        print(f"JSON:  {json_path} (exists: {json_path.exists()})")
        print("=" * 60)

    except Exception as exc:
        print(f"Error processing reports: {exc}")
        try:
            print("Attempting to generate reports with fallback...")
            generate_html([], {"passed": 0, "failed": 0, "skipped": 0}, 0, html_path, project_name, markers)
            generate_excel([], excel_path, project_name)
        except Exception as fallback_exc:
            print(f"Failed to generate reports: {fallback_exc}")
            return 1

    if exit_code != 0:
        print("Some tests failed")

    return exit_code


def main():
    args = parse_args()
    if getattr(args, "test", None):
        os.environ["MAVEN_TEST"] = args.test
    project_names = []
    if args.projects:
        project_names = [p.strip() for p in args.projects.split(",") if p.strip()]

    cli_project = _resolve_project_name(args.project)
    if cli_project:
        project_names.append(cli_project)

    if not project_names:
        project_names = [None]

    marker_expr = _resolve_markers(args.markers)

    overall_exit_code = 0
    for project_name in project_names:
        exit_code = _run_report_for_target(
            project_name, args.report_dir, args.skip_pytest, marker_expr
        )
        overall_exit_code = max(overall_exit_code, exit_code)

    return overall_exit_code


if __name__ == "__main__":
    raise SystemExit(main())
